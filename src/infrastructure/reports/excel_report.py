"""
Excel report generator for financial statements.

Produces a formatted .xlsx workbook with:
  - Sheet 1: Estado de Resultados (Income Statement)
  - Sheet 2: Balance General (Balance Sheet)
  - Sheet 3: Detalle de Movimientos (transaction detail, optional)

Uses openpyxl — install via: pip install openpyxl
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from src.domain.models.financial_statement import BalanceSheet, IncomeStatement

# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # Dark navy
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6") # Blue
_SUBTOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")  # Light blue
_POSITIVE_FONT = Font(bold=True, color="1A7A1A")         # Green
_NEGATIVE_FONT = Font(bold=True, color="C00000")          # Red
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")

_CURRENCY_FMT = '#,##0.00'
_PCT_FMT = '0.00%'


def _fmt_currency(ws, row: int, col: int, value: Decimal) -> None:
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = _CURRENCY_FMT
    cell.alignment = Alignment(horizontal="right")
    if value < 0:
        cell.font = _NEGATIVE_FONT


def _header_row(ws, row: int, *labels: str, fill=_HEADER_FILL) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def generate_income_statement_sheet(
    wb: openpyxl.Workbook,
    stmt: IncomeStatement,
) -> None:
    ws = wb.create_sheet("Estado de Resultados")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = f"Estado de Resultados — {stmt.period_label}"
    title.font = _TITLE_FONT
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Moneda: {stmt.currency}"
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    _header_row(ws, row, "Cuenta", "Código", "Monto")
    row += 1

    def section(title: str, lines, total: Decimal, fill=_SUBTOTAL_FILL) -> None:
        nonlocal row
        # Section header
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=f"  {title.upper()}")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _SUBHEADER_FILL
        row += 1
        for line in lines:
            ws.cell(row=row, column=1, value=f"    {line.account_name}").alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2, value=line.account_code)
            _fmt_currency(ws, row, 3, line.amount)
            row += 1
        # Subtotal
        cell_t = ws.cell(row=row, column=1, value=f"  TOTAL {title.upper()}")
        cell_t.font = Font(bold=True)
        cell_t.fill = fill
        _fmt_currency(ws, row, 3, total)
        ws.cell(row=row, column=3).fill = fill
        row += 2

    section("Ingresos", stmt.revenue_lines, stmt.total_revenue)
    section("Costo de Ventas", stmt.cogs_lines, stmt.total_cogs)

    # Gross profit
    ws.cell(row=row, column=1, value="UTILIDAD BRUTA").font = Font(bold=True, size=11)
    _fmt_currency(ws, row, 3, stmt.gross_profit)
    row += 2

    section("Gastos Operacionales", stmt.expense_lines, stmt.total_expenses)

    # Net income
    ws.merge_cells(f"A{row}:B{row}")
    net_cell = ws.cell(row=row, column=1, value="UTILIDAD NETA DEL PERÍODO")
    net_cell.font = Font(bold=True, size=12, color="FFFFFF")
    net_cell.fill = _HEADER_FILL
    _fmt_currency(ws, row, 3, stmt.net_income)
    ws.cell(row=row, column=3).fill = _HEADER_FILL
    ws.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF")

    _set_column_widths(ws, {1: 45, 2: 12, 3: 18})


def generate_balance_sheet_sheet(
    wb: openpyxl.Workbook,
    bs: BalanceSheet,
) -> None:
    ws = wb.create_sheet("Balance General")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = f"Balance General — {bs.period_label}"
    title.font = _TITLE_FONT
    title.alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Moneda: {bs.currency}  |  Ecuación: {'✓ CUADRA' if bs.is_balanced else '✗ NO CUADRA'}"
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    _header_row(ws, row, "Cuenta", "Código", "Monto")
    row += 1

    def section(title: str, sec, fill=_SUBTOTAL_FILL) -> None:
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        hdr = ws.cell(row=row, column=1, value=f"  {title.upper()}")
        hdr.font = Font(bold=True, color="FFFFFF")
        hdr.fill = _SUBHEADER_FILL
        row += 1
        for line in sec.lines:
            ws.cell(row=row, column=1, value=f"    {line.account_name}")
            ws.cell(row=row, column=2, value=line.account_code)
            _fmt_currency(ws, row, 3, line.amount)
            row += 1
        tot = ws.cell(row=row, column=1, value=f"  TOTAL {title.upper()}")
        tot.font = Font(bold=True)
        tot.fill = fill
        _fmt_currency(ws, row, 3, sec.total)
        ws.cell(row=row, column=3).fill = fill
        row += 2

    section("Activos", bs.assets)
    section("Pasivos", bs.liabilities)

    # Equity + retained earnings
    ws.merge_cells(f"A{row}:C{row}")
    hdr = ws.cell(row=row, column=1, value="  PATRIMONIO")
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = _SUBHEADER_FILL
    row += 1
    for line in bs.equity.lines:
        ws.cell(row=row, column=1, value=f"    {line.account_name}")
        ws.cell(row=row, column=2, value=line.account_code)
        _fmt_currency(ws, row, 3, line.amount)
        row += 1
    ws.cell(row=row, column=1, value="    Utilidades Retenidas Períodos Anteriores")
    _fmt_currency(ws, row, 3, bs.retained_earnings)
    row += 1
    ws.cell(row=row, column=1, value="    Utilidad del Período Actual")
    _fmt_currency(ws, row, 3, bs.current_period_net_income)
    row += 1
    total_eq = bs.equity.total + bs.retained_earnings + bs.current_period_net_income
    tot = ws.cell(row=row, column=1, value="  TOTAL PATRIMONIO")
    tot.font = Font(bold=True)
    tot.fill = _SUBTOTAL_FILL
    _fmt_currency(ws, row, 3, total_eq)
    ws.cell(row=row, column=3).fill = _SUBTOTAL_FILL
    row += 2

    # Final balance check
    ws.merge_cells(f"A{row}:B{row}")
    chk = ws.cell(row=row, column=1, value="TOTAL PASIVOS + PATRIMONIO")
    chk.font = Font(bold=True, size=12, color="FFFFFF")
    chk.fill = _HEADER_FILL
    _fmt_currency(ws, row, 3, bs.total_liabilities_and_equity)
    ws.cell(row=row, column=3).fill = _HEADER_FILL
    ws.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF")

    _set_column_widths(ws, {1: 45, 2: 12, 3: 18})


def generate_excel_report(
    income_stmt: IncomeStatement,
    balance_sheet: BalanceSheet,
    output_path: Path,
) -> Path:
    """
    Generate a formatted .xlsx workbook with both financial statements.

    Args:
        income_stmt:   IncomeStatement domain entity
        balance_sheet: BalanceSheet domain entity
        output_path:   Where to save the .xlsx file

    Returns:
        Path to the saved file
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    generate_income_statement_sheet(wb, income_stmt)
    generate_balance_sheet_sheet(wb, balance_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
