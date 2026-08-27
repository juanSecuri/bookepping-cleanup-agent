"""Export period financial statements to openpyxl workbook ($0) — 4 sheets."""
from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.use_cases.emit_period_reports import MONTH_KEYS, PeriodFinancialBundle

HEADER_FILL = PatternFill("solid", fgColor="1A4032")
HEADER_FONT = Font(bold=True, color="E8DCC8")
TOTAL_FONT = Font(bold=True)
NEG_FONT = Font(color="F87171")
TITLE_FONT = Font(bold=True, size=14, color="1A4032")


def _money_cell(ws, row: int, col: int, value: float | None) -> None:
    v = float(value or 0)
    cell = ws.cell(row=row, column=col, value=v)
    cell.number_format = '"$"#,##0.00'
    if v < 0:
        cell.font = NEG_FONT


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _append_pnl_section(ws, title: str, items: list, months: list[str], start_row: int) -> int:
    r = start_row
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
    r += 1
    headers = ["Código", "Cuenta", *[m for m in months], "Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    _style_header_row(ws, r, len(headers))
    r += 1
    section_total = 0.0
    for item in items:
        by_m = item.get("byMonth") or {}
        ws.cell(row=r, column=1, value=item.get("code"))
        ws.cell(row=r, column=2, value=item.get("name"))
        for i, mk in enumerate(months, 3):
            _money_cell(ws, r, i, float(by_m.get(mk) or 0))
        amt = float(item.get("amount") or 0)
        _money_cell(ws, r, 2 + len(months) + 1, amt)
        section_total += amt
        r += 1
    ws.cell(row=r, column=2, value=f"TOTAL {title}").font = TOTAL_FONT
    total_col = 2 + len(months) + 1
    _money_cell(ws, r, total_col, section_total)
    ws.cell(row=r, column=total_col).font = TOTAL_FONT
    return r + 2


def bundle_to_xlsx_bytes(
    bundle: PeriodFinancialBundle,
    transactions: list[Any] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    months = list(bundle.pnl.get("months") or MONTH_KEYS)

    # ── P&L ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "P&L"
    ws.append(["LedgerAI — Estado de resultados", bundle.period_label])
    ws["A1"].font = TITLE_FONT
    ws.append(["Motor", bundle.engine, "Txs verificadas", bundle.transaction_count])
    ws.append([])
    r = 4
    r = _append_pnl_section(ws, "INGRESOS", bundle.pnl.get("revenueItems") or [], months, r)
    r = _append_pnl_section(ws, "COSTO DE VENTAS", bundle.pnl.get("cogsItems") or [], months, r)
    r = _append_pnl_section(ws, "GASTOS OPERATIVOS", bundle.pnl.get("expenseItems") or [], months, r)
    ws.cell(row=r, column=2, value="UTILIDAD NETA").font = TOTAL_FONT
    net = float(bundle.pnl.get("netIncome") or bundle.pnl.get("net_income") or 0)
    total_col = 2 + len(months) + 1
    _money_cell(ws, r, total_col, net)
    ws.cell(row=r, column=total_col).font = TOTAL_FONT

    # ── Balance ────────────────────────────────────────────────────────────
    bs = wb.create_sheet("Balance")
    bal = bundle.balance_sheet or {}
    bs.append(["LedgerAI — Balance general", bundle.period_label])
    bs["A1"].font = TITLE_FONT
    bs.append(["Cuadra", bal.get("balanced"), "Ecuación", bal.get("equation")])
    bs.append([])
    headers_b = ["Sección", "Código", "Cuenta", "Saldo Inicial", "Débitos", "Créditos", "Saldo Final"]
    bs.append(headers_b)
    _style_header_row(bs, 4, len(headers_b))
    row_i = 5
    for section, key in (("ACTIVOS", "assets"), ("PASIVOS", "liabilities"), ("PATRIMONIO", "equity")):
        for item in bal.get(key) or []:
            bs.cell(row=row_i, column=1, value=section)
            bs.cell(row=row_i, column=2, value=item.get("code"))
            bs.cell(row=row_i, column=3, value=item.get("name"))
            _money_cell(bs, row_i, 4, float(item.get("opening") or 0))
            _money_cell(bs, row_i, 5, float(item.get("debits") or 0))
            _money_cell(bs, row_i, 6, float(item.get("credits") or 0))
            _money_cell(bs, row_i, 7, float(item.get("closing") or item.get("amount") or 0))
            row_i += 1
        total_key = {
            "assets": "totalAssets",
            "liabilities": "totalLiabilities",
            "equity": "totalEquity",
        }[key]
        bs.cell(row=row_i, column=3, value=f"TOTAL {section}").font = TOTAL_FONT
        _money_cell(bs, row_i, 7, float(bal.get(total_key) or 0))
        bs.cell(row=row_i, column=7).font = TOTAL_FONT
        row_i += 2
    bs.cell(row=row_i, column=1, value="Verificación A = P + E")
    bs.cell(row=row_i, column=2, value=bool(bal.get("balanced")))
    bs.cell(row=row_i, column=3, value=bal.get("equation"))

    # ── Cash flow ──────────────────────────────────────────────────────────
    cf = wb.create_sheet("Cash Flow")
    flow = bundle.cash_flow or {}
    detail = bundle.cash_flow_detail or {}
    cf.append(["LedgerAI — Cash flow O/I/F", bundle.period_label])
    cf["A1"].font = TITLE_FONT
    cf.append([])
    cf.append(["Tipo", "Código", "Cuenta", "Monto", "# txs"])
    _style_header_row(cf, 3, 5)
    row_i = 4
    for label, key, sub_key in (
        ("OPERATIVO", "operating", "operatingSubtotal"),
        ("INVERSIÓN", "investing", "investingSubtotal"),
        ("FINANCIACIÓN", "financing", "financingSubtotal"),
    ):
        lines = detail.get(key) or []
        if not lines:
            cf.cell(row=row_i, column=1, value=label)
            _money_cell(
                cf,
                row_i,
                4,
                float(detail.get(sub_key) or (flow.get(key) or {}).get("net") or 0),
            )
            row_i += 1
        else:
            for item in lines:
                cf.cell(row=row_i, column=1, value=label)
                cf.cell(row=row_i, column=2, value=item.get("code"))
                cf.cell(row=row_i, column=3, value=item.get("name"))
                _money_cell(cf, row_i, 4, float(item.get("amount") or 0))
                cf.cell(row=row_i, column=5, value=item.get("txCount"))
                row_i += 1
        cf.cell(row=row_i, column=3, value=f"SUBTOTAL {label}").font = TOTAL_FONT
        _money_cell(
            cf,
            row_i,
            4,
            float(detail.get(sub_key) or (flow.get(key) or {}).get("net") or 0),
        )
        cf.cell(row=row_i, column=4).font = TOTAL_FONT
        row_i += 2
    cf.cell(row=row_i, column=3, value="FLUJO NETO TOTAL").font = TOTAL_FONT
    _money_cell(cf, row_i, 4, float(detail.get("netTotal") or flow.get("netChange") or 0))
    cf.cell(row=row_i, column=4).font = TOTAL_FONT

    # ── Transacciones ──────────────────────────────────────────────────────
    tx = wb.create_sheet("Transacciones")
    tx.append(["LedgerAI — Transacciones del periodo", bundle.period_label])
    tx["A1"].font = TITLE_FONT
    tx.append([])
    tx.append(["Fecha", "Descripción", "Cuenta", "Código", "Tipo", "Monto", "Estado"])
    _style_header_row(tx, 3, 7)
    row_i = 4
    for t in transactions or []:
        d = str(getattr(t, "transaction_date", "") or "")
        if bundle.date_from and d and d < bundle.date_from:
            continue
        if bundle.date_to and d and d > bundle.date_to:
            continue
        status = getattr(t, "status", None)
        status_s = status.value if hasattr(status, "value") else str(status or "")
        if status_s not in ("verified", "closed"):
            continue
        tx_type = getattr(t, "transaction_type", None)
        type_s = tx_type.value if hasattr(tx_type, "value") else str(tx_type or "")
        tx.cell(row=row_i, column=1, value=d)
        tx.cell(row=row_i, column=2, value=getattr(t, "description", "") or "")
        tx.cell(row=row_i, column=3, value=getattr(t, "chart_of_accounts_name", "") or "")
        tx.cell(row=row_i, column=4, value=getattr(t, "chart_of_accounts_code", "") or "")
        tx.cell(row=row_i, column=5, value=type_s)
        _money_cell(tx, row_i, 6, float(getattr(t, "amount", 0) or 0))
        tx.cell(row=row_i, column=7, value=status_s)
        row_i += 1
    if row_i == 4:
        tx.cell(row=4, column=1, value="(Sin txs verificadas en el periodo)")

    for sheet in wb.worksheets:
        for col in range(1, 14):
            sheet.column_dimensions[get_column_letter(col)].width = 14
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 28

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
